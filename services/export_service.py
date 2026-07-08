"""
Excel 내보내기 서비스 — [M1.F8]

분석 결과를 openpyxl Workbook으로 변환하여 bytes로 반환한다.
라우트 레이어에서 Flask send_file()로 스트리밍 다운로드에 활용한다.
"""
import io
import logging
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

# CC 등급별 배경색 (hex, ARGB에서 FF 프리픽스 없이) — [M1.AC 8.1], [PR-7.2] 색상 토큰
_GRADE_FILL = {
    '낮음': PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid'),
    '중간': PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid'),
    '높음': PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid'),
    '매우높음': PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid'),
}

# God Class 탐지 임계값 — [M1.AC 2.5]
_GOD_CLASS_FUNCTION_COUNT = 20
_GOD_CLASS_AVG_CC = 8


def _collect_all_functions(analysis_results: dict) -> list:
    """전체 프로젝트 함수 목록을 단일 리스트로 합산."""
    functions = []
    for result in analysis_results.values():
        functions.extend(result.get('complexity', []))
    return functions


def _collect_all_sp_calls(analysis_results: dict) -> list:
    """전체 프로젝트 SP 호출 목록을 단일 리스트로 합산."""
    sp_calls = []
    for result in analysis_results.values():
        sp_calls.extend(result.get('sp_calls', []))
    return sp_calls


def _classify_class_entry(project: str, class_name: str, ccs: list) -> Optional[dict]:
    """클래스 CC 목록으로 God Class 판정 — 함수 수/평균 CC 임계값 검사 [M1.AC 2.5]."""
    count = len(ccs)
    avg_cc = sum(ccs) / count
    reasons = []
    if count >= _GOD_CLASS_FUNCTION_COUNT:
        reasons.append(f'함수 수 {count}개')
    if avg_cc >= _GOD_CLASS_AVG_CC:
        reasons.append(f'평균 CC {avg_cc:.1f}')
    if not reasons:
        return None
    return {
        'class_name': class_name,
        'project': project,
        'function_count': count,
        'avg_cc': round(avg_cc, 1),
        'reason': ', '.join(reasons),
    }


def _detect_god_classes(all_functions: list) -> list:
    """God Class 탐지 — 함수 수 >= 20 또는 평균 CC >= 8 [M1.AC 2.5]"""
    # (project, class_name) 기준 함수 그룹화
    class_groups: dict = defaultdict(list)
    for fn in all_functions:
        class_name = fn.get('class_name', '')
        if not class_name:
            continue
        key = (fn.get('project', ''), class_name)
        class_groups[key].append(fn.get('cc', 0))

    god_classes = []
    for (project, class_name), ccs in class_groups.items():
        entry = _classify_class_entry(project, class_name, ccs)
        if entry:
            god_classes.append(entry)
    return god_classes


def _write_sheet1(ws, all_functions: list) -> None:
    """시트 1: 전체 함수 목록 — CC 등급 배경색(PatternFill) [M1.AC 8.1]"""
    ws.append(['프로젝트', '파일', '클래스', '함수명', 'CC', '등급', '라인 수'])
    for fn in all_functions:
        ws.append([
            fn.get('project', ''),
            fn.get('file_path', ''),
            fn.get('class_name', ''),
            fn.get('function_name', ''),
            fn.get('cc', 0),
            fn.get('grade', ''),
            fn.get('line_start', 0),
        ])
        grade = fn.get('grade', '')
        fill = _GRADE_FILL.get(grade)
        if fill:
            # CC 컬럼(5)과 등급 컬럼(6)에 배경색 적용
            for col in (5, 6):
                ws.cell(row=ws.max_row, column=col).fill = fill


def _write_sheet2(ws, all_functions: list) -> None:
    """시트 2: 위험 함수 목록 — CC >= 10, CC 내림차순 [M1.AC 8.2]"""
    ws.append(['프로젝트', '파일', '클래스', '함수명', 'CC', '등급', '라인 수'])
    high_risk = sorted(
        (fn for fn in all_functions if fn.get('cc', 0) >= 10),
        key=lambda x: x.get('cc', 0),
        reverse=True,
    )
    for fn in high_risk:
        ws.append([
            fn.get('project', ''),
            fn.get('file_path', ''),
            fn.get('class_name', ''),
            fn.get('function_name', ''),
            fn.get('cc', 0),
            fn.get('grade', ''),
            fn.get('line_start', 0),
        ])


def _write_sheet3(ws, god_classes: list) -> None:
    """시트 3: God Class 목록 [M1.AC 8.3]"""
    ws.append(['클래스명', '프로젝트', '함수 수', '평균 CC', '판정 이유'])
    for gc in god_classes:
        ws.append([
            gc['class_name'],
            gc['project'],
            gc['function_count'],
            gc['avg_cc'],
            gc['reason'],
        ])


def _write_sheet4(ws, sp_calls: list) -> None:
    """시트 4: SP 사용 현황 [M1.AC 8.4]"""
    ws.append(['SP명', '사용 프로젝트', '파일 경로', '라인 번호', '호출 메서드'])
    for sp in sp_calls:
        ws.append([
            sp.get('sp_name', ''),
            sp.get('project', ''),
            sp.get('file_path', ''),
            sp.get('line_no', 0),
            sp.get('method_name', ''),
        ])


def _write_sheet5(ws, analysis_results: dict, god_classes: list, sp_calls: list) -> None:
    """시트 5: 프로젝트별 요약 통계 [M1.AC 8.5]"""
    ws.append(['프로젝트', '파일 수', '함수 수', '평균 CC', '최대 CC',
               '위험 함수 수', 'God Class 수', 'SP 수'])
    for proj_name, result in analysis_results.items():
        functions = result.get('complexity', [])
        proj_sp_count = sum(1 for sp in sp_calls if sp.get('project') == proj_name)
        proj_gc_count = sum(1 for gc in god_classes if gc['project'] == proj_name)
        files = {fn.get('file_path') for fn in functions}
        ccs = [fn.get('cc', 0) for fn in functions]
        avg_cc = round(sum(ccs) / len(ccs), 1) if ccs else 0.0
        max_cc = max(ccs) if ccs else 0
        high_risk = sum(1 for cc in ccs if cc >= 10)
        ws.append([
            proj_name,
            len(files),
            len(functions),
            avg_cc,
            max_cc,
            high_risk,
            proj_gc_count,
            proj_sp_count,
        ])


def _write_all_sheets(
    wb, all_functions: list, all_sp_calls: list, god_classes: list, analysis_results: dict
) -> None:
    """Workbook에 시트 1~5를 순서대로 작성 — [M1.AC 8.1]~[M1.AC 8.5]"""
    ws1 = wb.active
    ws1.title = '전체 함수 목록'
    _write_sheet1(ws1, all_functions)
    _write_sheet2(wb.create_sheet('위험 함수 목록'), all_functions)
    _write_sheet3(wb.create_sheet('God Class 목록'), god_classes)
    _write_sheet4(wb.create_sheet('SP 사용 현황'), all_sp_calls)
    _write_sheet5(wb.create_sheet('프로젝트별 요약'), analysis_results, god_classes, all_sp_calls)


def generate_excel_report(analysis_results: dict) -> bytes:
    """분석 결과를 Excel(xlsx) bytes로 변환 — [M1.AC 8.1]~[M1.AC 8.6]

    반환된 bytes를 Flask send_file()에 직접 전달하여 스트리밍 다운로드 가능.
    파일명 형식(PGAnalyzer_Report_YYYYMMDD_HHMMSS.xlsx)은 routes 레이어에서 결정.
    """
    wb = openpyxl.Workbook()
    all_functions = _collect_all_functions(analysis_results)
    all_sp_calls = _collect_all_sp_calls(analysis_results)
    god_classes = _detect_god_classes(all_functions)
    _write_all_sheets(wb, all_functions, all_sp_calls, god_classes, analysis_results)
    buf = io.BytesIO()
    wb.save(buf)
    logger.info('export_service: Excel 생성 완료 | projects=%d | functions=%d | sp_calls=%d',
                len(analysis_results), len(all_functions), len(all_sp_calls))
    return buf.getvalue()
