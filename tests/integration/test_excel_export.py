"""
[M1.IT-3] Happy — Excel 5개 시트 무결성

Ref: PRD Section 5-2 [M1.IT-3]
커버 AC: [M1.AC 8.1]~[M1.AC 8.6]

전제조건: 분석 완료 상태 (preloaded_cache fixture로 캐시 사전 로드)
"""
import io

import openpyxl
import pytest

# 실제 export_service가 생성하는 5개 시트명 [M1.AC 8.1]~[M1.AC 8.5]
EXPECTED_SHEETS = ['전체 함수 목록', '위험 함수 목록', 'God Class 목록', 'SP 사용 현황', '프로젝트별 요약']


@pytest.mark.happy
class TestExcelFiveSheets:
    """[M1.IT-3] Excel 내보내기 5개 시트 무결성 통합 테스트."""

    def test_excel_five_sheets(self, client, preloaded_cache):
        """
        GET /api/export/excel → openpyxl 파싱 → 5개 시트 검증.

        검증:
        1. HTTP 200, Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        2. openpyxl로 파일 파싱 성공
        3. 시트 5개 존재: 전체 함수 목록, 위험 함수 목록, God Class 목록, SP 사용 현황, 프로젝트별 요약
        4. 각 시트 헤더 행 정합성 (헤더 행 존재)
        5. 데이터 행 수 ≥ 1 (분석 결과와 일치)
        """
        # 1. GET /api/export/excel
        resp = client.get('/api/export/excel')
        assert resp.status_code == 200, f'Excel 응답 오류: {resp.status_code}'

        # Content-Type 검증
        content_type = resp.content_type
        assert 'spreadsheetml' in content_type, (
            f'Content-Type이 xlsx가 아닙니다: {content_type}'
        )

        # 2. openpyxl 파싱 성공
        excel_bytes = resp.data
        assert len(excel_bytes) > 0, 'Excel 파일이 비어 있습니다'
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))

        # 3. 시트 5개 존재 검증
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 5, (
            f'시트 수가 5개가 아닙니다. 실제: {sheet_names}'
        )
        for expected in EXPECTED_SHEETS:
            assert expected in sheet_names, (
                f'시트 "{expected}"가 없습니다. 실제 시트: {sheet_names}'
            )

        # 4. 각 시트 헤더 행 정합성 — 첫 번째 행에 헤더 존재
        for sheet_name in EXPECTED_SHEETS:
            ws = wb[sheet_name]
            first_row = [cell.value for cell in ws[1]]
            non_null_headers = [v for v in first_row if v is not None]
            assert len(non_null_headers) > 0, (
                f'시트 "{sheet_name}"의 헤더 행이 비어 있습니다'
            )

        # 5. '전체 함수 목록' 시트 데이터 행 수 ≥ 1 (preloaded_cache 결과 1개)
        ws_all = wb['전체 함수 목록']
        data_rows = [
            row for row in ws_all.iter_rows(min_row=2)
            if any(cell.value is not None for cell in row)
        ]
        assert len(data_rows) >= 1, (
            '전체 함수 목록 시트에 데이터 행이 없습니다 (헤더 제외)'
        )

        # 'SP 사용 현황' 시트 데이터 행 수 ≥ 1
        ws_sp = wb['SP 사용 현황']
        data_rows_sp = [
            row for row in ws_sp.iter_rows(min_row=2)
            if any(cell.value is not None for cell in row)
        ]
        assert len(data_rows_sp) >= 1, (
            'SP 사용 현황 시트에 데이터 행이 없습니다 (헤더 제외)'
        )

    def test_excel_export_fails_without_cache(self, client):
        """캐시 없을 때 Excel 내보내기 → 404 반환."""
        # clear_analysis_cache autouse fixture가 캐시를 비운 상태
        resp = client.get('/api/export/excel')
        assert resp.status_code == 404
        data = resp.get_json()
        assert 'error' in data
