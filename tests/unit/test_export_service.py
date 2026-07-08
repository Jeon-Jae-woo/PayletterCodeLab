"""
T3 > T3.8 — export_service 단위 테스트

[M1.F8]  Excel 내보내기
[M1.AC 8.1] 시트 1: 전체 함수 목록 (CC 등급 배경색 포함)
[M1.AC 8.2] 시트 2: 위험 함수 목록 (CC >= 10, 내림차순)
[M1.AC 8.3] 시트 3: God Class 목록
[M1.AC 8.4] 시트 4: SP 사용 현황
[M1.AC 8.5] 시트 5: 프로젝트별 요약 통계
[M1.AC 8.6] 파일명 형식, bytes 반환
"""
import io
import os
import sys

import openpyxl
import pytest

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)


# ---------------------------------------------------------------------------
# 공통 픽스처
# ---------------------------------------------------------------------------

def _make_function(name='Func', class_name='MyClass', project='ProjA',
                   file_path='/p/A.cs', cc=3, grade='낮음', line_start=10):
    return {
        'function_name': name,
        'class_name': class_name,
        'project': project,
        'file_path': file_path,
        'cc': cc,
        'grade': grade,
        'line_start': line_start,
        'category': 'normal',
    }


def _make_sp_call(sp_name='UP_GetPay', project='ProjA',
                  file_path='/p/A.cs', line_no=20, method_name='Execute'):
    return {
        'sp_name': sp_name,
        'project': project,
        'file_path': file_path,
        'line_no': line_no,
        'method_name': method_name,
        'class_name': 'MyClass',
        'category': 'payment',
    }


@pytest.fixture
def simple_results():
    """단일 프로젝트, 함수 3개, SP 1개 포함 최소 분석 결과."""
    return {
        'ProjA': {
            'complexity': [
                _make_function('Method1', cc=3, grade='낮음'),
                _make_function('Method2', cc=7, grade='중간'),
                _make_function('RiskyMethod', cc=12, grade='높음'),
            ],
            'sp_calls': [_make_sp_call()],
            'call_graph': {'nodes': [], 'edges': []},
            'dependency_graph': {'nodes': [], 'edges': []},
        }
    }


@pytest.fixture
def god_class_results():
    """God Class 탐지 조건(함수 수 >= 20) 충족 데이터."""
    functions = [_make_function(f'M{i}', class_name='HugeClass', cc=3)
                 for i in range(20)]
    return {
        'ProjB': {
            'complexity': functions,
            'sp_calls': [],
            'call_graph': {'nodes': [], 'edges': []},
            'dependency_graph': {'nodes': [], 'edges': []},
        }
    }


def _load_workbook(result_bytes: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(result_bytes))


# ---------------------------------------------------------------------------
# [M1.AC 8.6] 반환 타입
# ---------------------------------------------------------------------------

class TestGenerateExcelReport:
    """generate_excel_report() 기본 계약"""

    def test_returns_bytes(self, simple_results):
        """분석 결과 → bytes 반환 [M1.AC 8.6]"""
        from services import export_service
        result = export_service.generate_excel_report(simple_results)
        assert isinstance(result, bytes), "반환 타입이 bytes여야 합니다"
        assert len(result) > 0, "빈 bytes가 반환되면 안 됩니다"

    def test_empty_results_returns_bytes(self):
        """빈 분석 결과에서도 bytes 반환"""
        from services import export_service
        result = export_service.generate_excel_report({})
        assert isinstance(result, bytes), "빈 결과도 bytes를 반환해야 합니다"

    def test_workbook_has_five_sheets(self, simple_results):
        """Excel 파일에 정확히 5개 시트가 있어야 함 [M1.AC 8.1]~[M1.AC 8.5]"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        assert len(wb.sheetnames) == 5, f"시트 수는 5여야 합니다: {wb.sheetnames}"

    def test_sheet_names(self, simple_results):
        """5개 시트 이름이 요구사항과 일치해야 함"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        expected = {'전체 함수 목록', '위험 함수 목록', 'God Class 목록', 'SP 사용 현황', '프로젝트별 요약'}
        assert set(wb.sheetnames) == expected, f"시트명 불일치: {wb.sheetnames}"


# ---------------------------------------------------------------------------
# [M1.AC 8.1] 시트 1: 전체 함수 목록
# ---------------------------------------------------------------------------

class TestSheet1AllFunctions:
    """시트 1 — 전체 함수 목록"""

    def test_sheet1_has_correct_headers(self, simple_results):
        """시트 1 첫 행이 올바른 헤더여야 함"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['전체 함수 목록']
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        assert '프로젝트' in headers, "헤더에 '프로젝트' 포함되어야 합니다"
        assert '함수명' in headers, "헤더에 '함수명' 포함되어야 합니다"
        assert 'CC' in headers, "헤더에 'CC' 포함되어야 합니다"
        assert '등급' in headers, "헤더에 '등급' 포함되어야 합니다"

    def test_sheet1_data_row_count(self, simple_results):
        """시트 1에 헤더 + 3개 데이터 행이 있어야 함"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['전체 함수 목록']
        assert ws.max_row == 4, f"헤더 1 + 데이터 3 = 4행이어야 합니다: {ws.max_row}"

    def test_sheet1_cc_value_in_row(self, simple_results):
        """시트 1 데이터 행에 CC 값이 포함되어야 함"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['전체 함수 목록']
        # 헤더에서 CC 컬럼 위치 파악
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        cc_col = header_row.index('CC') + 1
        cc_values = [ws.cell(r, cc_col).value for r in range(2, ws.max_row + 1)]
        assert 3 in cc_values, "CC 값 3이 포함되어야 합니다"
        assert 12 in cc_values, "CC 값 12가 포함되어야 합니다"


# ---------------------------------------------------------------------------
# [M1.AC 8.2] 시트 2: 위험 함수 목록 (CC >= 10, 내림차순)
# ---------------------------------------------------------------------------

class TestSheet2HighRiskFunctions:
    """시트 2 — 위험 함수 목록"""

    def test_sheet2_only_high_risk(self, simple_results):
        """CC < 10 함수는 시트 2에 포함되지 않아야 함 [M1.AC 8.2]"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['위험 함수 목록']
        # 헤더 제외 데이터 행 수 = 1 (cc=12짜리만 포함)
        assert ws.max_row == 2, f"헤더 1 + 위험 함수 1 = 2행이어야 합니다: {ws.max_row}"

    def test_sheet2_sorted_by_cc_descending(self):
        """CC 내림차순 정렬 확인 [M1.AC 8.2]"""
        from services import export_service
        results = {
            'P': {
                'complexity': [
                    _make_function('F1', cc=15, grade='매우높음'),
                    _make_function('F2', cc=10, grade='높음'),
                    _make_function('F3', cc=12, grade='높음'),
                    _make_function('F4', cc=3, grade='낮음'),
                ],
                'sp_calls': [],
                'call_graph': {'nodes': [], 'edges': []},
                'dependency_graph': {'nodes': [], 'edges': []},
            }
        }
        wb = _load_workbook(export_service.generate_excel_report(results))
        ws = wb['위험 함수 목록']
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        cc_col = header_row.index('CC') + 1
        cc_values = [ws.cell(r, cc_col).value for r in range(2, ws.max_row + 1)]
        assert cc_values == sorted(cc_values, reverse=True), "CC 내림차순 정렬이어야 합니다"


# ---------------------------------------------------------------------------
# [M1.AC 8.3] 시트 3: God Class 목록
# ---------------------------------------------------------------------------

class TestSheet3GodClass:
    """시트 3 — God Class 목록"""

    def test_sheet3_detects_by_function_count(self, god_class_results):
        """함수 수 >= 20이면 God Class로 탐지 [M1.AC 2.5]"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(god_class_results))
        ws = wb['God Class 목록']
        assert ws.max_row >= 2, "God Class가 1건 이상 탐지되어야 합니다"

    def test_sheet3_detects_by_avg_cc(self):
        """평균 CC >= 8이면 God Class로 탐지 [M1.AC 2.5]"""
        from services import export_service
        results = {
            'P': {
                'complexity': [_make_function(f'M{i}', class_name='HighCC', cc=9)
                               for i in range(5)],
                'sp_calls': [],
                'call_graph': {'nodes': [], 'edges': []},
                'dependency_graph': {'nodes': [], 'edges': []},
            }
        }
        wb = _load_workbook(export_service.generate_excel_report(results))
        ws = wb['God Class 목록']
        assert ws.max_row >= 2, "평균 CC 9로 God Class가 탐지되어야 합니다"

    def test_sheet3_low_cc_not_god_class(self):
        """CC 낮고 함수 수 19개 이하면 God Class 아님"""
        from services import export_service
        results = {
            'P': {
                'complexity': [_make_function(f'M{i}', class_name='Normal', cc=3)
                               for i in range(5)],
                'sp_calls': [],
                'call_graph': {'nodes': [], 'edges': []},
                'dependency_graph': {'nodes': [], 'edges': []},
            }
        }
        wb = _load_workbook(export_service.generate_excel_report(results))
        ws = wb['God Class 목록']
        assert ws.max_row == 1, "헤더만 있어야 합니다 (God Class 없음)"


# ---------------------------------------------------------------------------
# [M1.AC 8.4] 시트 4: SP 사용 현황
# ---------------------------------------------------------------------------

class TestSheet4SpUsage:
    """시트 4 — SP 사용 현황"""

    def test_sheet4_sp_row_count(self, simple_results):
        """시트 4에 헤더 + SP 1건이 있어야 함 [M1.AC 8.4]"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['SP 사용 현황']
        assert ws.max_row == 2, f"헤더 1 + SP 1 = 2행이어야 합니다: {ws.max_row}"

    def test_sheet4_sp_name_in_data(self, simple_results):
        """시트 4에 SP명이 포함되어야 함"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['SP 사용 현황']
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        sp_col = header_row.index('SP명') + 1
        sp_names = [ws.cell(r, sp_col).value for r in range(2, ws.max_row + 1)]
        assert 'UP_GetPay' in sp_names, "SP명 'UP_GetPay'가 포함되어야 합니다"


# ---------------------------------------------------------------------------
# [M1.AC 8.5] 시트 5: 프로젝트별 요약 통계
# ---------------------------------------------------------------------------

class TestSheet5ProjectSummary:
    """시트 5 — 프로젝트별 요약 통계"""

    def test_sheet5_project_row_count(self, simple_results):
        """시트 5에 헤더 + 프로젝트 1건이 있어야 함 [M1.AC 8.5]"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['프로젝트별 요약']
        assert ws.max_row == 2, f"헤더 1 + 프로젝트 1 = 2행이어야 합니다: {ws.max_row}"

    def test_sheet5_high_risk_count(self, simple_results):
        """시트 5 위험 함수 수 = 1 (cc=12짜리 1개)"""
        from services import export_service
        wb = _load_workbook(export_service.generate_excel_report(simple_results))
        ws = wb['프로젝트별 요약']
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        risk_col = header_row.index('위험 함수 수') + 1
        assert ws.cell(2, risk_col).value == 1, "위험 함수 수는 1이어야 합니다"
